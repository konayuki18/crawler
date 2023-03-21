import os
import pickle as pk
from openpyxl import Workbook, load_workbook

wb = load_workbook('shopee.xlsx')
new_wb = Workbook()
ws = wb.active
new_ws = new_wb.active

# 設定行跟列的最大值
max_row = ws.max_row
max_column = ws.max_column

# 繁體轉簡體 簡體轉繁體 函式
class ZhhanzMan:
    def __init__(self):
        self.pkl_path = self.get_pkl_path()
        self.check_pkl()
        with open(self.pkl_path, 'rb') as pkl:
            self.dict = pk.load(pkl)

    def check_pkl(self):
        if not os.path.exists(self.pkl_path):
            import sys
            sys.path.append(self.get_dirpath())
            from php_conv import conv
            conv()

    def get_dirpath(self):
        pardir = os.path.abspath(__file__)
        return os.path.dirname(pardir)

    def get_pkl_path(self):
        pardir = self.get_dirpath()
        return os.path.join(pardir, 'zhhanz.pkl')

    def max_match(self, sent, d):
        for n in d['length']:
            seg = sent[:n]
            if d['dict'][n].get(seg, None):
                return d['dict'][n].get(seg)
            n -= 1
        return sent[0]

    def s2t(self, sent):
        return self.trans_s2t(sent)

    def t2s(self, sent):
        return self.trans_t2s(sent)

    # 簡體轉繁體
    def trans_s2t(self, sent):
        return self.trans(sent, self.dict['s2t'])

    # 繁體轉簡體
    def trans_t2s(self, sent):
        return self.trans(sent, self.dict['t2s'])

    def trans(self, sent, d):
        rtn = list()
        idx = 0
        while idx < len(sent):
            r = self.max_match(sent[idx:], d)
            idx += len(r)
            rtn.append(r)
        return ''.join(rtn)

if __name__ == '__main__':
    zm = ZhhanzMan()
    
    text_str = list()
    text_int = []
    
    for i in range(1, ws.max_row+1):
        for j in range(1, ws.max_column+1):
            # 判斷如果是文字就存在text_str裡
            if type(ws.cell(i, j).value) == str:
                text_str.append(ws.cell(i, j).value)
            # 判斷如果是數字就存在text_int裡
            elif type(ws.cell(i, j).value) == int:
                text_int.append(ws.cell(i, j).value)
        product = []
        # 呼叫trans_t2s()轉成簡體
        for a in text_str:
            product.append(zm.trans_t2s(a))
        # 數字直接存入
        for b in text_int:
            product.append(b)
        new_ws.append(product)
        text_int.clear()
        text_str.clear()
    
    new_wb.save('simplified_shopee.xlsx')
